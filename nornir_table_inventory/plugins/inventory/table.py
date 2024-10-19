import csv
import logging
from math import isnan
from typing import Any, Dict, List

# import pandas as pd
from nornir.core.inventory import (
    Inventory,
    Groups,
    Host,
    Hosts,
    Defaults,
    ConnectionOptions,
)
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


def _empty(x: Any):
    """Checks if x is a NaN (not a number) or None/empty string"""
    return x is None or (isinstance(x, float) and isnan(x)) or x == ""


def _get_connection_options(data: Dict[str, Any]) -> Dict[str, ConnectionOptions]:
    cp = {}
    for cn, c in data.items():
        cp[cn] = ConnectionOptions(
            hostname=c.get("hostname"),
            port=c.get("port"),
            username=c.get("username"),
            password=c.get("password"),
            platform=c.get("platform"),
            extras=c.get("extras"),
        )
    return cp


def _get_host_data(data: Dict[str, Any]) -> Dict[str, Any]:
    no_data_fields = ["name", "hostname", "port", "username", "password", "platform"]
    resp_data = {}
    netmiko_prefix = "netmiko_"
    for k, v in data.items():
        if (k not in no_data_fields) and (netmiko_prefix not in k):
            resp_data[k] = v if not _empty(v) else None
    return resp_data


def _get_host_netmiko_options(data: Dict[str, Any]) -> Dict[str, Any]:
    extra_opts = {}
    netmiko_options = {"netmiko": {"extras": {}}}
    """:cvar
    conn_timeout=5,
        auth_timeout=None,  # Timeout to wait for authentication response
        banner_timeout=15,  # Timeout to wait for the banner to be presented (post TCP-connect)
        # Other timeouts
        blocking_timeout=20,  # Read blocking timeout
        timeout=100,  # TCP connect timeout | overloaded to read-loop timeout
        session_timeout=60,  # Used for locking/sharing the connection
    
    
    """
    int_keys = "timeout conn_timeout auth_timeout banner_timeout blocking_timeout session_timeout".split()
    bool_keys = "fast_cli".split()
    netmiko_prefix = "netmiko_"
    for k, v in data.items():
        if netmiko_prefix in k:
            new_k = k.replace(netmiko_prefix, "")

            if new_k in int_keys:
                extra_opts[new_k] = int(v)
            elif new_k in bool_keys:
                if str(v).lower() in ["0", "false", "none"]:
                    extra_opts[new_k] = False
                else:
                    extra_opts[new_k] = True
            else:
                # if the value is nan,convert it to None
                if _empty(v):
                    extra_opts[new_k] = None
                else:
                    extra_opts[new_k] = v

    if extra_opts:
        netmiko_options["netmiko"]["extras"] = extra_opts
        return _get_connection_options(netmiko_options)
    else:
        return {}


def _get_host_obj(data: Dict[str, Any]) -> Host:
    # get keypoint data and convert to string or int
    name = data.get("name")
    hostname = data.get("hostname")
    port = data.get("port", 22)
    username = data.get("username")
    password = data.get("password")
    platform = data.get("platform")
    if name:
        name = str(name)
    if hostname:
        hostname = str(hostname) if not _empty(hostname) else None
    if port:
        port = int(port) if not _empty(port) else None
    if username:
        username = str(username) if not _empty(username) else None
    if password:
        password = str(password) if not _empty(password) else None
    if platform:
        platform = str(platform) if not _empty(platform) else None

    return Host(
        name=name,
        hostname=hostname,
        port=port,
        username=username,
        password=password,
        platform=platform,
        data=_get_host_data(data),
        groups=None,
        defaults={},
        connection_options=_get_host_netmiko_options(data),
    )


class FlatDataInventory:
    def __init__(self, data: List[Dict]) -> None:
        self.hosts_list = data

    def load(self) -> Inventory:
        defaults = Defaults()
        groups = Groups()
        hosts = Hosts()

        for host_dict in self.hosts_list:
            if not _empty(host_dict["name"]):
                hosts[host_dict["name"]] = _get_host_obj(host_dict)
            else:
                logger.error(f"HOST name is empty for data : {host_dict}")
                raise Exception("HOST name must not be empty")

        return Inventory(hosts=hosts, groups=groups, defaults=defaults)


class CSVInventory(FlatDataInventory):
    def __init__(self, csv_file: str = "inventory.csv") -> None:
        data = []
        with open(csv_file, mode="r", encoding="utf8") as f:
            for i in csv.DictReader(f):
                data.append(i)


class ExcelInventory(FlatDataInventory):
    """read in excel file and convert to inventory"""

    def __init__(self, excel_file: str = "inventory.xlsx") -> None:
        # self.hosts_list = []

        items: list[dict] = []
        # read in excel file
        workbook = load_workbook(excel_file)
        sheet: Worksheet = workbook.active
        headers = [cell.value for cell in sheet[1]]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            # print(row)
            item = {}
            for idx, cell_value in enumerate(row):
                # we skip the cell if the header is empty / None
                if headers[idx] is None or headers[idx] == "":
                    # print(f"WARNING: Empty header found in row - skipping cell value")
                    pass
                else:
                    item[headers[idx]] = cell_value
            items.append(item)
        # print(f"DEBUG: {items}")
        super().__init__(data=items)


if __name__ == "__main__":
    ...
